"""
Serviço de Exportação de Relatórios de Famílias em Excel.

Gera arquivos XLSX com resumos estatísticos e lista detalhada de famílias.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.services.familia_stats import FamiliaStatsService
from apps.core.models import Configuracao


class FamiliaExportService:
    """
    Serviço para exportar relatórios de famílias em Excel.
    
    Gera planilha com:
    - Aba 1: Resumo estatístico por categoria
    - Aba 2: Lista detalhada de famílias (todas ou filtradas por categoria)
    """
    
    # Estilos
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    CATEGORIAS_LABELS = {
        'maes_solo': '👩 Mães Solo (sem cônjuge)',
        'unipessoa': '👤 Famílias Unipessoais',
        'casal_sem_filho': '👫 Casal sem Filhos',
        '2': '👨‍👩‍👧 2 filhos',
        '3': '👨‍👩‍👧 3 filhos',
        '4': '👨‍👩‍👧 4 filhos',
        '5+': '👨‍👩‍👧 5 ou mais filhos',
        'todas': '📋 Todas as Famílias',
    }
    
    def __init__(self, stats_service: FamiliaStatsService):
        """
        Inicializa o serviço de exportação.
        
        Args:
            stats_service: Instância de FamiliaStatsService com filtros aplicados
        """
        self.stats = stats_service
    
    def _apply_header_style(self, cell):
        """Aplica estilo de cabeçalho a uma célula."""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.THIN_BORDER
    
    def _apply_cell_style(self, cell, center=False):
        """Aplica estilo padrão a uma célula."""
        cell.border = self.THIN_BORDER
        if center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_column_width(self, ws):
        """Ajusta largura das colunas automaticamente."""
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value and not isinstance(cell, type(None)):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_resumo_sheet(self, wb: Workbook):
        """Cria aba de resumo estatístico."""
        ws = wb.active
        ws.title = "Resumo"
        
        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = "Relatório de Composição Familiar"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Info do filtro
        ws.merge_cells('A2:E2')
        info = f"Lote: {self.stats.import_batch.imported_at.strftime('%d/%m/%Y') if self.stats.import_batch else 'Todos'}"
        if self.stats.filtros.get('bairro'):
            info += f" | Bairro: {self.stats.filtros['bairro']}"
        ws['A2'] = info
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = ['Tipo de Composição', 'Total', 'Aprovados', 'Reprovados', '% Aprovação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            self._apply_header_style(cell)
        
        # Dados
        dados = [
            ('maes_solo', self.stats.get_maes_solo()),
            ('unipessoa', self.stats.get_unipessoa()),
            ('casal_sem_filho', self.stats.get_casal_sem_filho()),
        ]
        
        # Adicionar categorias de filhos
        filhos = self.stats.get_filhos_quantitativos()
        for cat in ['2', '3', '4', '5+']:
            dados.append((cat, filhos.get(cat, {'total': 0, 'aprovados': 0, 'reprovados': 0, 'percentual_aprovacao': 0})))
        
        row = 5
        for cat_key, cat_data in dados:
            ws.cell(row=row, column=1, value=self.CATEGORIAS_LABELS.get(cat_key, cat_key))
            ws.cell(row=row, column=2, value=cat_data['total'])
            ws.cell(row=row, column=3, value=cat_data['aprovados'])
            ws.cell(row=row, column=4, value=cat_data['reprovados'])
            ws.cell(row=row, column=5, value=f"{cat_data['percentual_aprovacao']}%")
            
            for col in range(1, 6):
                self._apply_cell_style(ws.cell(row=row, column=col), center=(col > 1))
            row += 1
        
        # Seção por bairro
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "Distribuição por Bairro"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers_bairro = ['Bairro', 'Total', 'Aprovados', 'Reprovados', '% Aprovação']
        for col, header in enumerate(headers_bairro, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)
        row += 1
        
        por_bairro = self.stats.get_por_bairro()
        for bairro, dados_bairro in sorted(por_bairro.items()):
            ws.cell(row=row, column=1, value=bairro)
            ws.cell(row=row, column=2, value=dados_bairro['total'])
            ws.cell(row=row, column=3, value=dados_bairro['aprovados'])
            ws.cell(row=row, column=4, value=dados_bairro['reprovados'])
            ws.cell(row=row, column=5, value=f"{dados_bairro['percentual_aprovacao']}%")
            
            for col in range(1, 6):
                self._apply_cell_style(ws.cell(row=row, column=col), center=(col > 1))
            row += 1
        
        self._auto_column_width(ws)
    
    def _create_familias_sheet(self, wb: Workbook, categoria: str = 'todas'):
        """
        Cria aba com lista detalhada de famílias.
        
        Args:
            wb: Workbook
            categoria: Categoria para filtrar ('todas', 'maes_solo', etc.)
        """
        ws = wb.create_sheet(title="Famílias")
        
        # Título
        cat_label = self.CATEGORIAS_LABELS.get(categoria, 'Famílias')
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Lista de Famílias - {cat_label}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = [
            'Código Familiar',
            'Responsável Familiar',
            'CPF',
            'NIS',
            'Endereço',
            'Bairro',
            'Renda Média',
            'Status Validação'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)
        
        # Dados
        familias = self.stats.get_familias_para_exportacao(categoria)
        
        row = 4
        for familia in familias:
            # Buscar responsável familiar
            rf = familia.membros.filter(cod_parentesco_rf_pessoa=1).first()
            
            # Determinar status
            if getattr(familia, 'has_aprovada', False):
                status = 'Aprovado'
            elif getattr(familia, 'has_reprovada', False):
                status = 'Reprovado'
            else:
                status = 'Pendente'
            
            # Montar endereço
            endereco_parts = [familia.nom_logradouro_fam]
            if familia.num_logradouro_fam:
                endereco_parts.append(f"nº {familia.num_logradouro_fam}")
            endereco = ', '.join(filter(None, endereco_parts)) or '-'
            
            ws.cell(row=row, column=1, value=familia.cod_familiar_fam)
            ws.cell(row=row, column=2, value=rf.nom_pessoa if rf else '-')
            ws.cell(row=row, column=3, value=rf.num_cpf_pessoa if rf else '-')
            ws.cell(row=row, column=4, value=rf.num_nis_pessoa_atual if rf else '-')
            ws.cell(row=row, column=5, value=endereco)
            ws.cell(row=row, column=6, value=familia.nom_localidade_fam or '-')
            ws.cell(row=row, column=7, value=float(familia.vlr_renda_media_fam or 0))
            ws.cell(row=row, column=8, value=status)
            
            for col in range(1, 9):
                self._apply_cell_style(ws.cell(row=row, column=col))
            
            row += 1
        
        # Adicionar total
        ws.cell(row=row + 1, column=1, value=f"Total: {row - 4} famílias")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        
        self._auto_column_width(ws)
    
    def _get_nota_minima(self):
        """Retorna a nota mínima para aprovação configurada no sistema."""
        config = Configuracao.objects.first()
        return config.pontuacao_minima_aprovacao if config else 50
    
    def _create_validados_sheet(self, wb: Workbook, status_filtro: str, titulo_aba: str):
        """
        Cria aba com lista detalhada de famílias aprovadas ou reprovadas,
        incluindo pontuação.
        
        Args:
            wb: Workbook
            status_filtro: 'aprovado' ou 'reprovado'
            titulo_aba: Nome da aba
        """
        from apps.core.models import Validacao
        
        ws = wb.create_sheet(title=titulo_aba)
        
        nota_minima = self._get_nota_minima()
        
        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = f"Lista de Famílias {titulo_aba}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Info nota mínima
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Nota mínima para aprovação: {nota_minima} pontos"
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = [
            'Código Familiar',
            'Responsável Familiar',
            'CPF',
            'NIS',
            'Endereço',
            'Bairro',
            'Renda Média',
            'Pontuação',
            'Nota Mínima',
            'Status'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            self._apply_header_style(cell)
        
        # Buscar famílias com validação no status especificado
        familias_qs = self.stats.get_familias_queryset()
        
        # Filtrar por status da validação
        familias_com_validacao = familias_qs.filter(
            validacoes__status=status_filtro
        ).distinct().prefetch_related(
            'membros',
            'validacoes'
        )
        
        row = 5
        for familia in familias_com_validacao:
            # Buscar responsável familiar
            rf = familia.membros.filter(cod_parentesco_rf_pessoa=1).first()
            
            # Buscar a validação com o status
            validacao = familia.validacoes.filter(status=status_filtro).first()
            pontuacao = validacao.pontuacao_total if validacao else 0
            
            # Montar endereço
            endereco_parts = [familia.nom_logradouro_fam]
            if familia.num_logradouro_fam:
                endereco_parts.append(f"nº {familia.num_logradouro_fam}")
            endereco = ', '.join(filter(None, endereco_parts)) or '-'
            
            ws.cell(row=row, column=1, value=familia.cod_familiar_fam)
            ws.cell(row=row, column=2, value=rf.nom_pessoa if rf else '-')
            ws.cell(row=row, column=3, value=rf.num_cpf_pessoa if rf else '-')
            ws.cell(row=row, column=4, value=rf.num_nis_pessoa_atual if rf else '-')
            ws.cell(row=row, column=5, value=endereco)
            ws.cell(row=row, column=6, value=familia.nom_localidade_fam or '-')
            ws.cell(row=row, column=7, value=float(familia.vlr_renda_media_fam or 0))
            ws.cell(row=row, column=8, value=pontuacao)
            ws.cell(row=row, column=9, value=nota_minima)
            ws.cell(row=row, column=10, value='Aprovado' if status_filtro == 'aprovado' else 'Reprovado')
            
            # Aplicar estilos
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                self._apply_cell_style(cell)
                
                # Destaque na pontuação
                if col == 8:
                    if pontuacao >= nota_minima:
                        cell.font = Font(color="006600", bold=True)  # Verde
                    else:
                        cell.font = Font(color="CC0000", bold=True)  # Vermelho
            
            row += 1
        
        # Adicionar total
        total_familias = row - 5
        ws.cell(row=row + 1, column=1, value=f"Total: {total_familias} famílias")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        
        # Estatísticas de pontuação
        if total_familias > 0:
            ws.cell(row=row + 2, column=1, value="Estatísticas de Pontuação:")
            ws.cell(row=row + 2, column=1).font = Font(bold=True)
            
            # Calcular estatísticas
            pontuacoes = [
                familia.validacoes.filter(status=status_filtro).first().pontuacao_total
                for familia in familias_com_validacao
                if familia.validacoes.filter(status=status_filtro).exists()
            ]
            
            if pontuacoes:
                ws.cell(row=row + 3, column=1, value=f"Média: {sum(pontuacoes) / len(pontuacoes):.1f} pontos")
                ws.cell(row=row + 4, column=1, value=f"Maior: {max(pontuacoes)} pontos")
                ws.cell(row=row + 5, column=1, value=f"Menor: {min(pontuacoes)} pontos")
        
        self._auto_column_width(ws)
    
    def export_to_excel(self, categoria: str = 'todas') -> BytesIO:
        """
        Gera arquivo Excel completo.
        
        Args:
            categoria: Categoria para exportar ('todas', 'maes_solo', etc.)
        
        Returns:
            BytesIO com o arquivo Excel
        """
        wb = Workbook()
        
        # Criar abas
        self._create_resumo_sheet(wb)
        self._create_familias_sheet(wb, categoria)
        self._create_validados_sheet(wb, 'aprovado', 'Aprovados')
        self._create_validados_sheet(wb, 'reprovado', 'Reprovados')
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def get_filename(self, categoria: str = 'todas') -> str:
        """Gera nome do arquivo baseado nos filtros."""
        from datetime import datetime
        
        parts = ['relatorio_familias']
        
        if categoria != 'todas':
            parts.append(categoria.replace('+', 'mais'))
        
        if self.stats.filtros.get('bairro'):
            bairro_slug = self.stats.filtros['bairro'].lower().replace(' ', '_')[:20]
            parts.append(bairro_slug)
        
        parts.append(datetime.now().strftime('%Y%m%d_%H%M'))
        
        return '_'.join(parts) + '.xlsx'
