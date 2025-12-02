"""
Serviço para geração de arquivos XLSX (openpyxl) no formato BSDI.
"""
import io
from datetime import datetime
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from apps.cecad.models import ImportBatch
from apps.core.models import Validacao


class BSDIExporter:
    """Exportador de lista de beneficiários aprovados para o Banco Social de Dona Inês."""
    
    # Informações da instituição (podem ser movidas para configurações futuramente)
    INSTITUICAO_NOME = "Banco Solidário de Dona Inês"
    ENTIDADE_NOME = "Prefeitura Municipal de Dona Inês"
    RESPONSAVEL_NOME = "Julhio Arthur de Araújo Rodrigues"
    RESPONSAVEL_TELEFONE = "(83) 98192-5590"
    RESPONSAVEL_EMAIL = "bancosolidario@pmdonaines.pb.gov.br"
    
    def __init__(self, import_batch=None):
        """
        Inicializa o exportador.
        
        Args:
            import_batch: ImportBatch do qual extrair beneficiários. 
                         Se None, usa a última importação.
        """
        self.import_batch = import_batch or self._get_ultimo_batch()
        
        if not self.import_batch:
            raise ValueError("Nenhum lote de importação encontrado para exportação")
    
    def _get_ultimo_batch(self):
        """Retorna o último batch de importação concluído."""
        return ImportBatch.objects.filter(
            status='completed'
        ).order_by('-imported_at').first()
    
    def _get_familias_aprovadas(self):
        """
        Retorna queryset de famílias aprovadas do batch.
        
        Critérios:
        - Pertence ao import_batch especificado
        - Possui validação com status 'aprovado'
        """
        return self.import_batch.familias.filter(
            validacoes__status='aprovado'
        ).select_related().prefetch_related('membros').distinct()
    
    def _extrair_dados_beneficiario(self, familia):
        """
        Extrai dados do Responsável Familiar para exportação.
        
        Args:
            familia: Instância de Familia
            
        Returns:
            dict com os dados formatados para o XLS
        """
        rf = familia.get_responsavel_familiar()
        
        if not rf:
            # Se não encontrar RF, tentar o primeiro membro
            rf = familia.membros.first()
        
        # Formatar telefone (placeholder - pode ser expandido)
        telefone = ""
        
        # Formatar data de nascimento
        data_nascimento = ""
        if rf and rf.dat_nasc_pessoa:
            data_nascimento = rf.dat_nasc_pessoa.strftime('%d/%m/%Y')
        
        # Formatar CPF
        cpf = ""
        if rf and rf.num_cpf_pessoa:
            cpf = rf.num_cpf_pessoa
        
        # Nome completo
        nome = rf.nom_pessoa if rf else ""
        
        # Email (placeholder)
        email = ""
        
        # Endereço
        cep = familia.num_cep_logradouro_fam or ""
        endereco = familia.nom_logradouro_fam or ""
        numero = familia.num_logradouro_fam or ""
        complemento = ""
        bairro = familia.nom_localidade_fam or ""
        cidade_uf = "Dona Inês / PB"  # Fixo para este município
        
        return {
            'telefone': telefone,
            'data_nascimento': data_nascimento,
            'cpf': cpf,
            'nome': nome,
            'email': email,
            'cep': cep,
            'endereco': endereco,
            'numero': numero,
            'complemento': complemento,
            'bairro': bairro,
            'cidade_uf': cidade_uf,
        }
    
    def gerar_arquivo(self):
        """
        Gera o arquivo XLSX no formato BSDI.
        
        Returns:
            tuple: (ContentFile, nome_arquivo, total_beneficiarios)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Planilha1'

        # Dimensões de coluna baseadas na imagem
        # A=ORDEM, B=TELEFONE, C=DATA NASC, D=CPF, E=NOME, F=EMAIL, G=CEP, H=ENDEREÇO, I=NÚMERO, J=COMPLEMENTO, K=BAIRRO, L=CIDADE
        col_widths = [9, 13, 22, 13, 32, 24, 18, 20, 11, 16, 10, 12]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Alturas das linhas
        ws.row_dimensions[1].height = 16
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 16
        ws.row_dimensions[4].height = 16
        ws.row_dimensions[5].height = 8   # separador pequeno
        ws.row_dimensions[6].height = 16
        ws.row_dimensions[7].height = 28  # cabeçalhos (mais alto para wrap)

        # Estilos base
        thin = Side(style='thin', color='000000')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Cores conforme a imagem original
        fill_green_light = PatternFill('solid', fgColor='92D050')  # Verde claro mais saturado
        fill_green_dark = PatternFill('solid', fgColor='375623')   # Verde escuro para título

        font_bold = Font(bold=True, size=10)
        font_bold_small = Font(bold=True, size=9)
        font_normal = Font(size=9)
        font_white_bold = Font(bold=True, color='FFFFFF', size=10)
        font_red = Font(color='FF0000', size=9)  # Vermelho
        font_blue_link = Font(color='0563C1', underline='single', size=9)  # Azul link

        # ============================================================
        # Linha 1: Título principal - VERDE ESCURO
        # ============================================================
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1, value='RESPONSÁVEL PELA ABERTURA DE CONTAS DE BENEFICIÁRIOS')
        title_cell.font = font_white_bold
        title_cell.alignment = center
        title_cell.fill = fill_green_dark
        title_cell.border = border_all
        for col in range(2, 7):
            ws.cell(row=1, column=col).border = border_all

        # ============================================================
        # Linha 2: Banco Solidário | Valor: R$ | Nome da instituição: | Prefeitura
        # ============================================================
        # A2-B2: "Banco Solidário de Dona Inês" (mesclado, negrito)
        ws.merge_cells('A2:B2')
        a2 = ws.cell(row=2, column=1, value=self.INSTITUICAO_NOME)
        a2.font = font_bold_small
        a2.fill = fill_green_light
        a2.border = border_all
        a2.alignment = left_align
        ws.cell(row=2, column=2).border = border_all

        # C2: "Valor: R$"
        c2 = ws.cell(row=2, column=3, value='Valor: R$')
        c2.font = font_normal
        c2.fill = fill_green_light
        c2.border = border_all
        c2.alignment = left_align

        # D2: "Nome da instituição:"
        d2 = ws.cell(row=2, column=4, value='Nome da instituição:')
        d2.font = font_normal
        d2.fill = fill_green_light
        d2.border = border_all
        d2.alignment = left_align

        # E2-L2: Valor da instituição em vermelho (mesclado)
        ws.merge_cells('E2:L2')
        inst = ws.cell(row=2, column=5, value=self.ENTIDADE_NOME)
        inst.font = font_red
        inst.alignment = left_align
        inst.fill = fill_green_light
        inst.border = border_all
        for col in range(6, 13):
            ws.cell(row=2, column=col).border = border_all

        # ============================================================
        # Linha 3: Responsável pelo cadastro de BENEFICIÁRIOS:
        # ============================================================
        # A3-D3: Label (mesclado)
        ws.merge_cells('A3:D3')
        lab_resp = ws.cell(row=3, column=1, value='Responsável pelo cadastro de BENEFICIÁRIOS:')
        lab_resp.font = font_normal
        lab_resp.fill = fill_green_light
        lab_resp.border = border_all
        lab_resp.alignment = left_align
        for col in range(2, 5):
            ws.cell(row=3, column=col).border = border_all

        # E3-L3: Nome do responsável em vermelho (mesclado)
        ws.merge_cells('E3:L3')
        resp = ws.cell(row=3, column=5, value=self.RESPONSAVEL_NOME)
        resp.font = font_red
        resp.alignment = left_align
        resp.fill = fill_green_light
        resp.border = border_all
        for col in range(6, 13):
            ws.cell(row=3, column=col).border = border_all

        # ============================================================
        # Linha 4: Telefone e Email
        # ============================================================
        # A4: "Telefone:"
        a4 = ws.cell(row=4, column=1, value='Telefone:')
        a4.font = font_normal
        a4.fill = fill_green_light
        a4.border = border_all
        a4.alignment = left_align

        # B4: Número do telefone
        b4 = ws.cell(row=4, column=2, value=self.RESPONSAVEL_TELEFONE)
        b4.font = font_normal
        b4.fill = fill_green_light
        b4.border = border_all
        b4.alignment = left_align

        # C4: "Email:"
        c4 = ws.cell(row=4, column=3, value='Email:')
        c4.font = font_normal
        c4.fill = fill_green_light
        c4.border = border_all
        c4.alignment = left_align

        # D4-L4: Email (mesclado) - azul/link
        ws.merge_cells('D4:L4')
        email_cell = ws.cell(row=4, column=4, value=self.RESPONSAVEL_EMAIL)
        email_cell.font = font_blue_link
        email_cell.fill = fill_green_light
        email_cell.border = border_all
        email_cell.alignment = left_align
        for col in range(5, 13):
            ws.cell(row=4, column=col).border = border_all

        # ============================================================
        # Linha 5: Separador (vazio, sem bordas)
        # ============================================================
        # Linha vazia de separação

        # ============================================================
        # Linha 6: Título da seção "CADASTRAMENTO PARA ABERTURA DE CONTAS DE BENEFICIÁRIOS"
        # ============================================================
        ws.merge_cells('A6:L6')
        sec = ws.cell(row=6, column=1, value='CADASTRAMENTO PARA ABERTURA DE CONTAS DE BENEFICIÁRIOS')
        sec.font = font_bold
        sec.alignment = Alignment(horizontal='center', vertical='center')
        sec.fill = fill_green_light
        sec.border = border_all
        for col in range(2, 13):
            ws.cell(row=6, column=col).border = border_all

        # ============================================================
        # Linha 7: Cabeçalhos das colunas
        # ============================================================
        headers = [
            'ORDEM',
            'TELEFONE',
            'DATA DE NASCIMENTO',
            'CPF',
            'NOME COMPLETO',
            'ENDEREÇO DE E-MAIL',
            'CEP DA RESIDÊNCIA',
            'ENDEREÇO',
            'NÚMERO DA CASA',
            'COMPLEMENTO',
            'BAIRRO',
            'CIDADE / UF',
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = font_bold_small
            cell.fill = fill_green_light
            cell.border = border_all
            cell.alignment = wrap_center

        # Freeze panes abaixo dos cabeçalhos
        ws.freeze_panes = ws['A8']

        # AutoFilter nos cabeçalhos
        ws.auto_filter.ref = "A7:L7"

        # ============================================================
        # Dados dos beneficiários (a partir da linha 8)
        # ============================================================
        familias = self._get_familias_aprovadas()

        row_idx = 8
        ordem = 1
        for familia in familias:
            dados = self._extrair_dados_beneficiario(familia)
            values = [
                ordem,
                dados['telefone'],
                dados['data_nascimento'],
                dados['cpf'],
                dados['nome'],
                dados['email'],
                dados['cep'],
                dados['endereco'],
                dados['numero'],
                dados['complemento'],
                dados['bairro'],
                dados['cidade_uf'],
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border_all
                cell.font = font_normal
                cell.alignment = center if col == 1 else left_align
            row_idx += 1
            ordem += 1

        total_beneficiarios = ordem - 1

        # Salvar em buffer como .xlsx
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f'lista_bsdi_{timestamp}.xlsx'
        content_file = ContentFile(buffer.read(), name=nome_arquivo)

        return content_file, nome_arquivo, total_beneficiarios
